import openpyxl
from Model import PriceItem

def read_price(price_file):
    wb = openpyxl.load_workbook(price_file)
    sheet = wb['Sheet1']
    result = list()
    row_tmp = 1
    column_tmp = 5
    while column_tmp < sheet.max_column:
        while row_tmp < sheet.max_row:
            if sheet.cell(row=row_tmp, column=column_tmp).value == None:
                price = sheet.cell(row=row_tmp, column=5).value
            else:
                price = sheet.cell(row=row_tmp, column=column_tmp).value
            target = sheet.cell(row=row_tmp, column=2).value
            project = sheet.cell(row=row_tmp, column=3).value
            method = sheet.cell(row=row_tmp, column=4).value
            custom = sheet.cell(row=2, column=column_tmp).value
            tmp = PriceItem(target, project, method, price, custom)
            result.append(tmp)
            row_tmp += 1
        row_tmp = 1
        column_tmp += 1
    return result

result1 = read_price('中科基因各个客户收费价格备案表20200426.xlsx')
print("result size:"+str(len(result1)))
for item in result1:
    print(str(item.custom) + "  " + str(item.target) + "  " + str(item.project) + "  " + str(item.method) + "  " + str(item.price))
