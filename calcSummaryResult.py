import openpyxl
import xlrd
from Model import ResultItem

def calcSummaryResult(data_file):
    result = list()
    wb = openpyxl.load_workbook(data_file, data_only=True)
    sheet = wb['检测金额统计总表']
    row_tmp = 3
    column_tmp = 3
    while column_tmp < sheet.max_column:
        while row_tmp < sheet.max_row:

            #记录每个客户，每个月的金额
            custom = sheet.cell(row=row_tmp, column=2).value
            month = xlrd.xldate_as_tuple(sheet.cell(row=1, column=column_tmp).value, 0)[1]
            value = sheet.cell(row=row_tmp, column=column_tmp).value
            total = sheet.cell(row=row_tmp, column=15).value
            item_tmp = ResultItem(custom, month, value, total)
            result.append(item_tmp)

            row_tmp += 1
        row_tmp = 1
        column_tmp += 1
    return result

result1 = calcSummaryResult('data1.xlsx')
print("result size:"+str(len(result1)))
for item in result1:
    print(str(item.custom) + "  " + str(item.month) + "  " + str(item.value) + "  " + str(item.total))